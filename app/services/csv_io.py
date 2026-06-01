import csv
import io
import re
import zipfile
from dataclasses import dataclass
from ipaddress import IPv4Address, IPv4Network, ip_address

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import pyzipper

from app.models import IPAddress, Project
from app.services.network import reserved_project_addresses

BASE_COLUMNS = ["ip", "hostname", "os", "type", "comment"]
EXPORT_COLUMNS = BASE_COLUMNS
XLSX_MAX_UNCOMPRESSED_BYTES = 20_000_000
FIELD_LIMITS = {
    "hostname": 255,
    "os": 120,
    "type": 120,
    "comment": 4000,
}


class CSVImportError(ValueError):
    pass


@dataclass(frozen=True)
class ImportedAsset:
    address: str
    hostname: str
    os: str
    asset_type: str
    comment: str


@dataclass(frozen=True)
class CSVImportResult:
    cidr: str
    rows: list[ImportedAsset]


def _decode_csv(content: bytes) -> str:
    for encoding in ("utf-8-sig", "cp1251"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise CSVImportError("Не удалось прочитать файл: используйте UTF-8 или Windows-1251")


def _validate_headers(headers: list[str], *, source: str) -> list[str]:
    if headers == BASE_COLUMNS:
        return BASE_COLUMNS
    expected = "ip;hostname;os;type;comment"
    raise CSVImportError(f"Неверный заголовок {source}. Ожидается: {expected}")


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _smallest_network(addresses: list[IPv4Address]) -> IPv4Network:
    first = min(int(address) for address in addresses)
    last = max(int(address) for address in addresses)
    prefix = 32 - (first ^ last).bit_length()
    mask = (0xFFFFFFFF << (32 - prefix)) & 0xFFFFFFFF if prefix else 0
    return IPv4Network((first & mask, prefix))


def _parse_import_rows(
    rows_with_numbers: list[tuple[int, dict[str, str]]],
    *,
    max_addresses: int,
) -> CSVImportResult:
    rows: list[ImportedAsset] = []
    parsed_addresses: list[IPv4Address] = []
    seen_addresses: dict[str, int] = {}

    for row_number, raw_row in rows_with_numbers:
        normalized = {key: _cell_text(raw_row.get(key)) for key in EXPORT_COLUMNS}
        if not any(normalized.values()):
            continue

        raw_ip = normalized["ip"]
        if not raw_ip:
            raise CSVImportError(f"Строка {row_number}: поле ip обязательно")

        try:
            parsed_ip = ip_address(raw_ip)
        except ValueError as exc:
            raise CSVImportError(f"Строка {row_number}: неверный IPv4-адрес") from exc

        if not isinstance(parsed_ip, IPv4Address):
            raise CSVImportError(f"Строка {row_number}: поддерживаются только IPv4-адреса")

        address = str(parsed_ip)
        if address in seen_addresses:
            raise CSVImportError(
                f"Строка {row_number}: IP {address} уже указан в строке {seen_addresses[address]}"
            )
        seen_addresses[address] = row_number

        for field_name, limit in FIELD_LIMITS.items():
            if len(normalized[field_name]) > limit:
                raise CSVImportError(f"Строка {row_number}: поле {field_name} длиннее {limit} символов")

        rows.append(
            ImportedAsset(
                address=address,
                hostname=normalized["hostname"],
                os=normalized["os"],
                asset_type=normalized["type"],
                comment=normalized["comment"],
            )
        )
        parsed_addresses.append(parsed_ip)

        if len(rows) > max_addresses:
            raise CSVImportError(f"CSV содержит больше строк, чем разрешенный лимит проекта: {max_addresses}")

    if not rows:
        raise CSVImportError("CSV не содержит строк с IP-адресами")

    network = _smallest_network(parsed_addresses)
    if network.num_addresses > max_addresses:
        raise CSVImportError(
            f"По импортированным IP получилась подсеть {network} на {network.num_addresses} адресов, "
            f"лимит: {max_addresses}"
        )

    reserved = reserved_project_addresses(str(network))
    reserved_hits = [row.address for row in rows if row.address in reserved]
    if reserved_hits:
        sample = ", ".join(reserved_hits[:5])
        raise CSVImportError(f"CSV содержит network/broadcast адреса для подсети {network}: {sample}")

    rows.sort(key=lambda item: int(ip_address(item.address)))
    return CSVImportResult(cidr=str(network), rows=rows)


def parse_assets_csv(content: bytes, *, max_addresses: int) -> CSVImportResult:
    text = _decode_csv(content)
    reader = csv.DictReader(io.StringIO(text, newline=""), delimiter=";")
    if reader.fieldnames is None:
        raise CSVImportError("CSV-файл пустой или не содержит заголовок")

    headers = [header.strip().lower() for header in reader.fieldnames]
    columns = _validate_headers(headers, source="CSV")

    rows_with_numbers: list[tuple[int, dict[str, str]]] = []
    for row_number, raw_row in enumerate(reader, start=2):
        if None in raw_row:
            raise CSVImportError(f"Строка {row_number}: слишком много столбцов")
        rows_with_numbers.append((row_number, {key: _cell_text(raw_row.get(key)) for key in columns}))

    return _parse_import_rows(rows_with_numbers, max_addresses=max_addresses)


def parse_assets_xlsx(content: bytes, *, max_addresses: int) -> CSVImportResult:
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            uncompressed_size = sum(item.file_size for item in archive.infolist())
    except zipfile.BadZipFile as exc:
        raise CSVImportError("Не удалось прочитать XLSX-файл") from exc
    if uncompressed_size > XLSX_MAX_UNCOMPRESSED_BYTES:
        raise CSVImportError("XLSX-файл слишком большой после распаковки")

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except (InvalidFileException, OSError, ValueError) as exc:
        raise CSVImportError("Не удалось прочитать XLSX-файл") from exc

    worksheet = workbook.worksheets[0]
    rows_iter = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as exc:
        raise CSVImportError("XLSX-файл пустой или не содержит заголовок") from exc

    headers = [_cell_text(item).lower() for item in header_row if _cell_text(item)]
    columns = _validate_headers(headers, source="XLSX")

    rows_with_numbers: list[tuple[int, dict[str, str]]] = []
    for row_number, values in enumerate(rows_iter, start=2):
        row_values = list(values[: len(columns)])
        if len(values) > len(columns) and any(_cell_text(item) for item in values[len(columns):]):
            raise CSVImportError(f"Строка {row_number}: слишком много столбцов")
        rows_with_numbers.append(
            (
                row_number,
                {column: _cell_text(row_values[index]) if index < len(row_values) else "" for index, column in enumerate(columns)},
            )
        )

    return _parse_import_rows(rows_with_numbers, max_addresses=max_addresses)


def parse_assets_file(content: bytes, *, filename: str, max_addresses: int) -> CSVImportResult:
    clean_filename = filename.lower().strip()
    if clean_filename.endswith(".xlsx"):
        return parse_assets_xlsx(content, max_addresses=max_addresses)
    if clean_filename.endswith(".csv"):
        return parse_assets_csv(content, max_addresses=max_addresses)
    raise CSVImportError("Поддерживаются только файлы .csv и .xlsx")


def render_project_csv(project: Project, ip_records: list[IPAddress]) -> str:
    output = io.StringIO(newline="")
    writer = csv.writer(output, delimiter=";", lineterminator="\n")
    writer.writerow(EXPORT_COLUMNS)
    for ip_record in ip_records:
        writer.writerow(
            [
                ip_record.address,
                ip_record.hostname,
                ip_record.os,
                ip_record.asset_type,
                ip_record.comment,
            ]
        )
    return output.getvalue()


def render_project_xlsx(project: Project, ip_records: list[IPAddress]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Assets"
    worksheet.append(EXPORT_COLUMNS)
    for ip_record in ip_records:
        worksheet.append(
            [
                ip_record.address,
                ip_record.hostname,
                ip_record.os,
                ip_record.asset_type,
                ip_record.comment,
            ]
        )

    widths = {
        "A": 18,
        "B": 24,
        "C": 22,
        "D": 18,
        "E": 42,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def csv_bytes(content: str) -> bytes:
    return ("\ufeff" + content).encode("utf-8")


def safe_export_name(value: str, *, suffix: str) -> str:
    clean = re.sub(r"[^A-Za-z0-9_.-]+", "_", value.strip()).strip("._")
    return f"{clean or 'export'}{suffix}"


def _build_aes_zip_archive(files: dict[str, bytes], password: str) -> bytes:
    buffer = io.BytesIO()
    with pyzipper.AESZipFile(
        buffer,
        "w",
        compression=zipfile.ZIP_DEFLATED,
        encryption=pyzipper.WZ_AES,
    ) as archive:
        archive.setpassword(password.encode("utf-8"))
        for filename, content in files.items():
            archive.writestr(filename, content)
    return buffer.getvalue()


def build_zip_archive(files: dict[str, bytes], *, password: str | None = None) -> bytes:
    if password:
        return _build_aes_zip_archive(files, password)

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for filename, content in files.items():
            archive.writestr(filename, content)

    return buffer.getvalue()
